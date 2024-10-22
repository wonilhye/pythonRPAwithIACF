import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment
import os  # os 모듈을 임포트하여 파일 열기

# 엑셀 파일 경로
file_path = 'C:\\IACFPYTHON\\codesample\\dev\\xlsx\\data.xlsx'
sheet_name = '과제데이터'

# 과제데이터 시트에서 데이터 읽기
df = pd.read_excel(file_path, sheet_name=sheet_name)

# 피벗 테이블 생성: 시작년월별로 월지급액의 합을 구함
pivot_table = pd.pivot_table(df, 
                             values='월지급액',  # 집계할 값
                             index='시작년월',   # 행 그룹화 기준
                             aggfunc='sum')      # 합계 함수 사용

# 인덱스를 열로 변환하여 새로운 DataFrame 생성 (엑셀에 쓰기 위함)
pivot_table_reset = pivot_table.reset_index()

# 월지급액의 총합 계산
total_amount = pivot_table['월지급액'].sum()

# 총합을 새로운 행으로 추가 ('합계'라는 텍스트와 함께)
total_row = pd.DataFrame({"시작년월": ["합계"], "월지급액": [total_amount]})
pivot_table_reset = pd.concat([pivot_table_reset, total_row], ignore_index=True)

# 기존 엑셀 파일을 열어 "append" 모드로 시트 추가
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # "월별금액현황" 시트가 존재하면 덮어쓰기
    pivot_table_reset.to_excel(writer, sheet_name='월별금액현황', index=False)

# 엑셀 파일 다시 열어서 수정
book = load_workbook(file_path)
sheet = book['월별금액현황']

# 월지급액 열을 정수 형식으로 지정하고 천 단위 쉼표 추가
for row in range(2, sheet.max_row + 1):  # 2번째 행부터 마지막 행까지
    cell = sheet.cell(row=row, column=2)  # 월지급액 열의 셀
    cell.number_format = '#,##0'  # 천 단위로 쉼표가 있는 정수 형식

# "월지급액" 열의 너비를 120픽셀로 설정
sheet.column_dimensions['A'].width = 15  # A열(월지급액)의 너비 설정
sheet.column_dimensions['B'].width = 15  # B열(월지급액)의 너비 설정

# A열 중앙 정렬
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):  # A열(첫 번째 열)만 중앙 정렬
    for cell in row:
        cell.alignment = Alignment(horizontal='center')  # 중앙 정렬

# 엑셀 파일 저장
book.save(file_path)

# 최종 파일 열기
os.startfile(file_path)  # 작업 완료 후 엑셀 파일 열기